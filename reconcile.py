import pandas as pd
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# ─── COLOURS ────────────────────────────────────────────────────────────────
BG = {
    'dk_blue':  '1F3864', 'md_blue':  '2E75B6', 'lt_blue':  'D6E4F0',
    'dk_green': '1E4620', 'md_green': '375623', 'lt_green': 'E2EFDA',
    'dk_red':   '7B0000', 'md_red':   'C00000', 'lt_red':   'FFE2E2', 'pink': 'FFD7D7',
    'yellow':   'FFF2CC', 'orange':   'FCE4D6', 'amber':    'FFC000',
    'grey':     'F2F2F2', 'mid_grey': 'BFBFBF', 'white':    'FFFFFF',
    'purple':   '4A235A', 'lt_purple':'F5EEF8',
}
FG = {
    'white': 'FFFFFF', 'black': '000000', 'md_red': 'C00000',
    'md_green': '375623', 'dk_red': '7B0000', 'md_blue': '2E75B6',
    'grey': '595959', 'amber': 'B8860B', 'purple': '4A235A',
    'lt_purple': '4A235A',
}

def _c(ws, row, col, val=None, bold=False, bg='white', fg='black',
        sz=10, ha='left', wrap=False, fmt=None, italic=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name='Arial', bold=bold, color=FG.get(fg, fg), size=sz, italic=italic)
    cell.fill = PatternFill('solid', fgColor=BG.get(bg, bg))
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    return cell

def _mr(ws, row, c1, c2, val=None, **kw):
    ws.merge_cells(f'{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}')
    _c(ws, row, c1, val, **kw)
    bg = kw.get('bg', 'white')
    for c in range(c1 + 1, c2 + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=BG.get(bg, 'FFFFFF'))

def _hdr(ws, row, cols, bg='md_blue'):
    for col, lbl in cols:
        cell = ws.cell(row=row, column=col, value=lbl)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cell.fill = PatternFill('solid', fgColor=BG[bg])
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 15

def fd(ts):
    try:
        return pd.Timestamp(ts).strftime('%d/%m/%Y') if pd.notna(ts) else ''
    except Exception:
        return ''


# ─── SAP EXPORT PARSER ───────────────────────────────────────────────────────
def parse_sap_export(filepath_or_buffer):
    """
    Reads a SAP customer line-item export (FBL5N style or the ALV export used here).
    Returns a normalised DataFrame with columns:
      assignment, doc_number, doc_type, doc_date, due_date,
      amount, clearing_doc, clearing_date, text, header_text
    """
    raw = pd.read_excel(filepath_or_buffer, sheet_name=0, header=0)
    raw.columns = [str(c).strip() for c in raw.columns]

    # Map common SAP column name variants
    col_map = {
        'Assignment':               'assignment',
        'Document Number':          'doc_number',
        'Document Type':            'doc_type',
        'Document Date':            'doc_date',
        'Net due date':             'due_date',
        'Amount in local currency': 'amount',
        'Clearing Document':        'clearing_doc',
        'Clearing date':            'clearing_date',
        'Text':                     'text',
        'Document Header Text':     'header_text',
        'Reference Key 1':          'ref_key1',
        'Reference Key 3':          'ref_key3',
    }
    rename = {k: v for k, v in col_map.items() if k in raw.columns}
    df = raw.rename(columns=rename)

    for col in ['doc_date', 'due_date', 'clearing_date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    if 'amount' in df.columns:
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)

    if 'assignment' in df.columns:
        df['assignment_str'] = df['assignment'].astype(str).str.strip()
    else:
        df['assignment_str'] = ''

    return df


# ─── REMITTANCE PARSER ───────────────────────────────────────────────────────
def _extract_refs(text):
    """Pull all 9954xxxxxx-style AB-InBev reference numbers from a string."""
    return re.findall(r'9954\d{6}', str(text))

def parse_remittance(filepath_or_buffer):
    """
    Reads a client remittance / payment advice file.
    Tries to find reference numbers and amounts in any layout.
    Returns a list of dicts: {ref, invoice_num, amount, status, raw_ref}
    """
    raw = pd.read_excel(filepath_or_buffer, sheet_name=0, header=None)

    items = []
    ref_col, amt_col, inv_col, status_col = None, None, None, None

    # Scan first 10 rows for a header
    for hdr_row_idx in range(min(10, len(raw))):
        row_vals = [str(v).strip().lower() for v in raw.iloc[hdr_row_idx]]
        if any('ref' in v or 'factuur' in v or 'invoice' in v for v in row_vals):
            for ci, v in enumerate(row_vals):
                if 'referentie' in v or 'reference' in v or 'ref' in v:
                    ref_col = ci
                if 'totaal' in v or 'bedrag' in v or 'amount' in v or 'total' in v:
                    amt_col = ci
                if 'factuur' in v or 'invoice' in v or 'nummer' in v:
                    inv_col = ci
                if 'status' in v:
                    status_col = ci
            if ref_col is not None:
                # Data starts after header
                for _, data_row in raw.iloc[hdr_row_idx + 1:].iterrows():
                    vals = data_row.tolist()
                    raw_ref = str(vals[ref_col]).strip() if ref_col < len(vals) else ''
                    refs = _extract_refs(raw_ref)
                    if not refs and raw_ref and raw_ref != 'nan':
                        refs = [raw_ref]
                    amt = vals[amt_col] if amt_col is not None and amt_col < len(vals) else None
                    try:
                        amt = float(str(amt).replace(',', '.').replace(' ', '')) if amt not in (None, '', 'nan') else None
                    except Exception:
                        amt = None
                    inv = str(vals[inv_col]).strip() if inv_col is not None and inv_col < len(vals) else ''
                    status = str(vals[status_col]).strip() if status_col is not None and status_col < len(vals) else ''
                    for ref in refs:
                        items.append({'ref': ref, 'invoice_num': inv, 'amount': amt,
                                      'status': status, 'raw_ref': raw_ref})
            break

    # Fallback: scan every cell for 9954xxxxxx patterns
    if not items:
        for _, row in raw.iterrows():
            for cell in row:
                refs = _extract_refs(cell)
                for ref in refs:
                    items.append({'ref': ref, 'invoice_num': '', 'amount': None,
                                  'status': '', 'raw_ref': str(cell)})

    return items


# ─── MAIN RECONCILIATION LOGIC ───────────────────────────────────────────────
def run_reconciliation(sap_file, remittance_file, payment_amount=None, payment_date=None):
    """
    Core reconciliation engine.
    Returns a dict with all findings ready for report generation.
    """
    sap = parse_sap_export(sap_file)
    remittance_items = parse_remittance(remittance_file)

    # ── SAP OPEN ITEMS (no clearing document) ──
    open_items = sap[sap['clearing_doc'].isna()].copy() if 'clearing_doc' in sap.columns else sap.copy()
    cleared_items = sap[sap['clearing_doc'].notna()].copy() if 'clearing_doc' in sap.columns else pd.DataFrame()

    open_rv = open_items[open_items.get('doc_type', pd.Series(dtype=str)) == 'RV'] if 'doc_type' in open_items.columns else open_items
    open_invoices = open_rv[open_rv['amount'] > 0] if 'amount' in open_rv.columns else open_rv
    open_credits = open_rv[open_rv['amount'] < 0] if 'amount' in open_rv.columns else pd.DataFrame()
    open_ru = open_items[open_items.get('doc_type', pd.Series(dtype=str)) == 'RU'] if 'doc_type' in open_items.columns else pd.DataFrame()

    # ── BUILD LOOKUP: assignment → open item ──
    open_ref_lookup = {}
    if 'assignment_str' in open_items.columns:
        for _, row in open_items.iterrows():
            ref = row['assignment_str']
            if ref and ref != 'nan':
                open_ref_lookup.setdefault(ref, []).append(row)

    # ── BUILD LOOKUP: assignment → cleared item ──
    cleared_ref_lookup = {}
    if 'assignment_str' in cleared_items.columns:
        for _, row in cleared_items.iterrows():
            ref = row['assignment_str']
            if ref and ref != 'nan':
                cleared_ref_lookup.setdefault(ref, []).append(row)

    # ── MATCH REMITTANCE ITEMS ──
    matched = []          # ref found open in SAP, amounts align
    not_found = []        # ref not found open in SAP at all
    already_cleared = []  # ref found but already has a clearing doc
    amount_diff = []      # ref found open but amounts differ
    remittance_refs = set()

    for item in remittance_items:
        ref = item['ref']
        remittance_refs.add(ref)
        rem_amt = item['amount']

        if ref in open_ref_lookup:
            sap_rows = open_ref_lookup[ref]
            sap_amt = sum(r['amount'] for r in sap_rows)
            diff = (rem_amt - sap_amt) if rem_amt is not None else None
            if diff is not None and abs(diff) > 1.0:
                amount_diff.append({**item, 'sap_amount': sap_amt, 'difference': diff,
                                     'sap_rows': sap_rows})
            else:
                matched.append({**item, 'sap_amount': sap_amt, 'sap_rows': sap_rows})
        elif ref in cleared_ref_lookup:
            cleared_rows = cleared_ref_lookup[ref]
            already_cleared.append({**item, 'sap_rows': cleared_rows,
                                     'cleared_by': cleared_rows[0].get('clearing_doc', ''),
                                     'cleared_date': cleared_rows[0].get('clearing_date', '')})
        else:
            not_found.append(item)

    # ── INVOICES IN SAP NOT IN REMITTANCE ──
    if 'due_date' in open_invoices.columns:
        cutoff = pd.Timestamp('2099-12-31')
        if payment_date:
            try:
                cutoff = pd.Timestamp(payment_date) + pd.DateOffset(days=30)
            except Exception:
                pass
        missing_from_remittance = open_invoices[
            ~open_invoices['assignment_str'].isin(remittance_refs) &
            (open_invoices['amount'] > 0)
        ]
    else:
        missing_from_remittance = pd.DataFrame()

    # ── TOTALS ──
    remittance_total = sum(i['amount'] for i in remittance_items if i['amount'] is not None)
    matched_total = sum(i['sap_amount'] for i in matched)
    already_cleared_total = sum(
        sum(r['amount'] for r in i['sap_rows']) for i in already_cleared
    )
    not_found_total = sum(i['amount'] for i in not_found if i['amount'] is not None)
    amount_diff_total = sum(i['difference'] for i in amount_diff if i['difference'] is not None)

    open_balance = open_items['amount'].sum() if 'amount' in open_items.columns else 0
    open_credits_total = open_credits['amount'].sum() if 'amount' in open_credits.columns and len(open_credits) > 0 else 0
    open_ru_total = open_ru['amount'].sum() if 'amount' in open_ru.columns and len(open_ru) > 0 else 0

    return {
        'matched': matched,
        'not_found': not_found,
        'already_cleared': already_cleared,
        'amount_diff': amount_diff,
        'missing_from_remittance': missing_from_remittance,
        'open_invoices': open_invoices,
        'open_credits': open_credits,
        'open_ru': open_ru,
        'remittance_total': remittance_total,
        'matched_total': matched_total,
        'already_cleared_total': already_cleared_total,
        'not_found_total': not_found_total,
        'amount_diff_total': amount_diff_total,
        'open_balance': open_balance,
        'open_credits_total': open_credits_total,
        'open_ru_total': open_ru_total,
        'payment_amount': payment_amount,
        'payment_date': payment_date,
        'remittance_items': remittance_items,
        'sap_df': sap,
    }


# ─── EXCEL REPORT BUILDER ────────────────────────────────────────────────────
def build_report(results, customer_name='Customer'):
    wb = openpyxl.Workbook()

    # ── SHEET 1: SUMMARY ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    for col, w in zip('ABCDEFGH', [4, 40, 4, 18, 4, 18, 4, 4]):
        ws.column_dimensions[col].width = w

    r = 1
    _mr(ws, r, 1, 8, 'REMITTANCE RECONCILIATION REPORT', bold=True, bg='dk_blue', fg='white', sz=16, ha='center')
    ws.row_dimensions[r].height = 40; r += 1
    _mr(ws, r, 1, 8, f'Customer: {customer_name}   ·   Payment: €{results["payment_amount"]:,.2f}' if results["payment_amount"] else f'Customer: {customer_name}',
        bg='md_blue', fg='white', sz=10, ha='center', italic=True)
    ws.row_dimensions[r].height = 20; r += 1
    ws.row_dimensions[r].height = 10; r += 1

    # Score card
    total_items = len(results['remittance_items'])
    n_matched = len(results['matched'])
    n_cleared = len(results['already_cleared'])
    n_missing = len(results['not_found'])
    n_diff = len(results['amount_diff'])

    _mr(ws, r, 1, 8, 'RECONCILIATION SUMMARY', bold=True, bg='dk_blue', fg='white', sz=11, ha='center')
    ws.row_dimensions[r].height = 24; r += 1

    summary_rows = [
        ('Payment amount on remittance', results['remittance_total'], 'lt_blue', False, 'black'),
        ('', None, 'white', False, 'black'),
        (f'✓  Invoices matched & open in SAP ({n_matched} items)', results['matched_total'], 'lt_green', False, 'md_green'),
        (f'⚠  Already cleared — potential double payment ({n_cleared} items)', results['already_cleared_total'], 'lt_red', False, 'md_red'),
        (f'✗  Not found in SAP at all ({n_missing} items)', results['not_found_total'], 'pink', False, 'md_red'),
        (f'△  Amount differences vs SAP ({n_diff} items)', results['amount_diff_total'], 'orange', False, 'black'),
        ('', None, 'white', False, 'black'),
        ('Open SAP balance (all invoice types)', results['open_balance'], 'lt_blue', False, 'black'),
        ('   of which: open credit notes available', results['open_credits_total'], 'lt_green', False, 'md_green'),
        ('   of which: open goods returns (RU)', results['open_ru_total'], 'lt_green', False, 'md_green'),
    ]
    for desc, amt, bg, bold, fg in summary_rows:
        if desc == '':
            _mr(ws, r, 1, 8, None, bg='white'); ws.row_dimensions[r].height = 6; r += 1; continue
        _mr(ws, r, 1, 5, desc, bold=bold, bg=bg, fg=fg, sz=10)
        if amt is not None:
            _c(ws, r, 6, amt, bold=bold, bg=bg, fg=fg, fmt='#,##0.00', ha='right', sz=10)
        else:
            _c(ws, r, 6, None, bg=bg)
        _c(ws, r, 7, None, bg=bg); _c(ws, r, 8, None, bg=bg)
        ws.row_dimensions[r].height = 20; r += 1

    # ── SHEET 2: MATCHED ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('✓ Matched')
    for col, w in zip('ABCDEF', [4, 22, 30, 16, 16, 4]):
        ws2.column_dimensions[col].width = w
    r2 = 1
    _mr(ws2, r2, 1, 6, f'✓  MATCHED INVOICES — Found open in SAP  ({n_matched} items)', bold=True, bg='md_green', fg='white', sz=11)
    ws2.row_dimensions[r2].height = 24; r2 += 1
    _mr(ws2, r2, 1, 6, 'These invoices are open in SAP and match the remittance. They are being paid correctly.', bg='lt_green', fg='md_green', sz=9, italic=True)
    ws2.row_dimensions[r2].height = 16; r2 += 1
    _hdr(ws2, r2, [(1, '#'), (2, 'Reference'), (3, 'Invoice # (remittance)'), (4, 'Remittance Amt (€)'), (5, 'SAP Amount (€)'), (6, '')])
    r2 += 1
    for idx, item in enumerate(results['matched'], 1):
        bg = 'lt_green' if idx % 2 == 0 else 'white'
        _c(ws2, r2, 1, idx, bg=bg, sz=8, ha='center')
        _c(ws2, r2, 2, item['ref'], bg=bg, sz=9)
        _c(ws2, r2, 3, item.get('invoice_num', ''), bg=bg, sz=9)
        _c(ws2, r2, 4, item['amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
        _c(ws2, r2, 5, item['sap_amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
        _c(ws2, r2, 6, None, bg=bg)
        ws2.row_dimensions[r2].height = 13; r2 += 1
    _mr(ws2, r2, 1, 3, 'TOTAL', bold=True, bg='md_green', fg='white', sz=10)
    _c(ws2, r2, 4, results['matched_total'], bold=True, bg='md_green', fg='white', fmt='#,##0.00', ha='right', sz=10)
    _c(ws2, r2, 5, None, bg='md_green'); _c(ws2, r2, 6, None, bg='md_green')
    ws2.row_dimensions[r2].height = 16

    # ── SHEET 3: ALREADY CLEARED (potential doubles) ───────────────────────────
    ws3 = wb.create_sheet('⚠ Already Cleared')
    for col, w in zip('ABCDEFG', [4, 22, 28, 16, 18, 22, 4]):
        ws3.column_dimensions[col].width = w
    r3 = 1
    _mr(ws3, r3, 1, 7, f'⚠  ALREADY CLEARED IN SAP — Potential Double Payments  ({n_cleared} items)', bold=True, bg='md_red', fg='white', sz=11)
    ws3.row_dimensions[r3].height = 24; r3 += 1
    _mr(ws3, r3, 1, 7, 'These items appear on the remittance but already have a clearing document in SAP. Investigate before processing — they may have been paid in a prior payment run.',
        bg='lt_red', fg='dk_red', sz=9, italic=True, wrap=True)
    ws3.row_dimensions[r3].height = 20; r3 += 1
    _hdr(ws3, r3, [(1, '#'), (2, 'Reference'), (3, 'Invoice # (remittance)'), (4, 'Amount (€)'), (5, 'Cleared Date in SAP'), (6, 'SAP Clearing Doc'), (7, '')])
    r3 += 1
    for idx, item in enumerate(results['already_cleared'], 1):
        _c(ws3, r3, 1, idx, bg='lt_red', sz=8, ha='center')
        _c(ws3, r3, 2, item['ref'], bg='lt_red', sz=9, bold=True)
        _c(ws3, r3, 3, item.get('invoice_num', ''), bg='lt_red', sz=9)
        _c(ws3, r3, 4, item['amount'], bg='lt_red', fmt='#,##0.00', ha='right', sz=9)
        _c(ws3, r3, 5, fd(item.get('cleared_date')), bg='lt_red', sz=9, ha='center')
        _c(ws3, r3, 6, str(item.get('cleared_by', '')), bg='lt_red', sz=9)
        _c(ws3, r3, 7, None, bg='lt_red')
        ws3.row_dimensions[r3].height = 14; r3 += 1

    # ── SHEET 4: NOT FOUND ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet('✗ Not Found in SAP')
    for col, w in zip('ABCDE', [4, 22, 30, 16, 4]):
        ws4.column_dimensions[col].width = w
    r4 = 1
    _mr(ws4, r4, 1, 5, f'✗  NOT FOUND IN SAP  ({n_missing} items)', bold=True, bg='purple', fg='white', sz=11)
    ws4.row_dimensions[r4].height = 24; r4 += 1
    _mr(ws4, r4, 1, 5, 'These references appear on the remittance but cannot be found anywhere in the SAP export. They may be under a different reference, not yet booked, or may not exist.',
        bg='lt_purple', fg='purple', sz=9, italic=True, wrap=True)
    ws4.row_dimensions[r4].height = 20; r4 += 1
    _hdr(ws4, r4, [(1, '#'), (2, 'Reference'), (3, 'Invoice # (remittance)'), (4, 'Amount (€)'), (5, '')])
    r4 += 1
    for idx, item in enumerate(results['not_found'], 1):
        bg = 'lt_purple' if idx % 2 == 0 else 'white'
        _c(ws4, r4, 1, idx, bg=bg, sz=8, ha='center')
        _c(ws4, r4, 2, item['ref'], bg=bg, sz=9)
        _c(ws4, r4, 3, item.get('invoice_num', ''), bg=bg, sz=9)
        _c(ws4, r4, 4, item['amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
        _c(ws4, r4, 5, None, bg=bg)
        ws4.row_dimensions[r4].height = 13; r4 += 1

    # ── SHEET 5: AMOUNT DIFFERENCES ───────────────────────────────────────────
    ws5 = wb.create_sheet('△ Amount Differences')
    for col, w in zip('ABCDEFG', [4, 22, 28, 16, 16, 16, 4]):
        ws5.column_dimensions[col].width = w
    r5 = 1
    _mr(ws5, r5, 1, 7, f'△  AMOUNT DIFFERENCES  ({n_diff} items)', bold=True, bg='amber', fg='white', sz=11)
    ws5.row_dimensions[r5].height = 24; r5 += 1
    _mr(ws5, r5, 1, 7, 'These references were found in SAP but the amounts differ from the remittance by more than €1.00. Review these carefully.',
        bg='yellow', fg='black', sz=9, italic=True)
    ws5.row_dimensions[r5].height = 16; r5 += 1
    _hdr(ws5, r5, [(1, '#'), (2, 'Reference'), (3, 'Invoice # (remittance)'), (4, 'Remittance Amt (€)'), (5, 'SAP Amount (€)'), (6, 'Difference (€)'), (7, '')])
    r5 += 1
    for idx, item in enumerate(results['amount_diff'], 1):
        bg = 'yellow' if idx % 2 == 0 else 'white'
        _c(ws5, r5, 1, idx, bg=bg, sz=8, ha='center')
        _c(ws5, r5, 2, item['ref'], bg=bg, sz=9)
        _c(ws5, r5, 3, item.get('invoice_num', ''), bg=bg, sz=9)
        _c(ws5, r5, 4, item['amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
        _c(ws5, r5, 5, item['sap_amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
        fg_diff = 'md_red' if item['difference'] < 0 else 'md_green'
        _c(ws5, r5, 6, item['difference'], bg=bg, fg=fg_diff, fmt='#,##0.00', ha='right', sz=9, bold=True)
        _c(ws5, r5, 7, None, bg=bg)
        ws5.row_dimensions[r5].height = 13; r5 += 1

    # ── SHEET 6: OPEN SAP ITEMS NOT ON REMITTANCE ─────────────────────────────
    if len(results['missing_from_remittance']) > 0:
        ws6 = wb.create_sheet('SAP Open — Not on Remittance')
        for col, w in zip('ABCDEF', [4, 22, 30, 14, 14, 4]):
            ws6.column_dimensions[col].width = w
        r6 = 1
        mfr = results['missing_from_remittance']
        _mr(ws6, r6, 1, 6, f'OPEN IN SAP — NOT INCLUDED ON REMITTANCE  ({len(mfr)} items  ·  €{mfr["amount"].sum():,.2f})',
            bold=True, bg='md_blue', fg='white', sz=11)
        ws6.row_dimensions[r6].height = 24; r6 += 1
        _mr(ws6, r6, 1, 6, 'These invoices are open in SAP but not mentioned on the remittance. The customer may have forgotten them or plans a separate payment.',
            bg='lt_blue', fg='md_blue', sz=9, italic=True, wrap=True)
        ws6.row_dimensions[r6].height = 16; r6 += 1
        _hdr(ws6, r6, [(1, '#'), (2, 'SAP Reference'), (3, 'Description'), (4, 'Invoice Date'), (5, 'Due Date'), (6, 'Amount (€)')])
        r6 += 1
        for idx, (_, row) in enumerate(mfr.sort_values('due_date').iterrows(), 1):
            bg = 'lt_blue' if idx % 2 == 0 else 'white'
            _c(ws6, r6, 1, idx, bg=bg, sz=8, ha='center')
            _c(ws6, r6, 2, row.get('assignment_str', ''), bg=bg, sz=9)
            _c(ws6, r6, 3, str(row.get('header_text', '')) if pd.notna(row.get('header_text')) else '', bg=bg, sz=9)
            _c(ws6, r6, 4, fd(row.get('doc_date')), bg=bg, sz=9, ha='center')
            _c(ws6, r6, 5, fd(row.get('due_date')), bg=bg, sz=9, ha='center')
            _c(ws6, r6, 6, row['amount'], bg=bg, fmt='#,##0.00', ha='right', sz=9)
            ws6.row_dimensions[r6].height = 13; r6 += 1
        _mr(ws6, r6, 1, 5, 'TOTAL', bold=True, bg='md_blue', fg='white', sz=10)
        _c(ws6, r6, 6, mfr['amount'].sum(), bold=True, bg='md_blue', fg='white', fmt='#,##0.00', ha='right', sz=10)
        ws6.row_dimensions[r6].height = 16

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
